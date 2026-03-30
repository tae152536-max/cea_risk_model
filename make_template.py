"""
make_template.py
----------------
Two public functions consumed by n_state_app.py:

    create_template_bytes(n_states, state_names) -> bytes
        Returns a formatted .xlsx template the user can download, fill in,
        and upload back.

    parse_template(file) -> (params_dict | None, error_str | None)
        Parses a filled-in template and returns the model parameters.
"""

import io
import numpy as np
import pandas as pd

# ─────────────────────────────────────────────────────────────────────────────
# TEMPLATE CREATION
# ─────────────────────────────────────────────────────────────────────────────

def create_template_bytes(n_states: int = 3, state_names: list = None) -> bytes:
    buf = io.BytesIO()
    _write_template(buf, n_states, state_names)
    buf.seek(0)
    return buf.read()


def _write_template(target, n_states: int = 3, state_names: list = None):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    if state_names is None:
        defaults = ["Well", "Sick", "Dead", "Severe", "Critical",
                    "State 6", "State 7", "State 8", "State 9", "State 10"]
        state_names = defaults[:n_states]
    while len(state_names) < n_states:
        state_names.append(f"State {len(state_names)+1}")
    state_names = list(state_names)[:n_states]

    wb = openpyxl.Workbook()

    # ── Shared styles ──────────────────────────────────────────────────────
    thin = Side(style="thin", color="94a3b8")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    NAVY   = PatternFill("solid", fgColor="1e3a5f")
    BLUE   = PatternFill("solid", fgColor="2563eb")
    GREEN  = PatternFill("solid", fgColor="dcfce7")
    YELLOW = PatternFill("solid", fgColor="fef9c3")
    WHITE  = PatternFill("solid", fgColor="FFFFFF")

    def style(cell, *, bold=False, color="000000", bg=None, italic=False,
              size=10, halign="left", num_fmt=None):
        cell.font      = Font(bold=bold, color=color, italic=italic, size=size)
        cell.alignment = Alignment(horizontal=halign, vertical="center",
                                   wrap_text=True)
        if bg:
            cell.fill = bg
        cell.border = brd
        if num_fmt:
            cell.number_format = num_fmt

    def title_row(ws, row, text, end_col):
        ws.merge_cells(
            start_row=row, start_column=1, end_row=row, end_column=end_col
        )
        c = ws.cell(row=row, column=1, value=text)
        style(c, bold=True, color="FFFFFF", bg=NAVY, size=12, halign="center")
        ws.row_dimensions[row].height = 22

    def section_row(ws, row, text, end_col):
        ws.merge_cells(
            start_row=row, start_column=1, end_row=row, end_column=end_col
        )
        c = ws.cell(row=row, column=1, value=text)
        style(c, bold=True, color="FFFFFF", bg=BLUE, size=10, halign="left")
        ws.row_dimensions[row].height = 18

    def inst_row(ws, row, text, end_col):
        ws.merge_cells(
            start_row=row, start_column=1, end_row=row, end_column=end_col
        )
        c = ws.cell(row=row, column=1, value=text)
        style(c, italic=True, color="1d4ed8", bg=YELLOW, size=9, halign="left")
        ws.row_dimensions[row].height = 16

    def kv(ws, row, key, value, num_fmt=None):
        k = ws.cell(row=row, column=1, value=key)
        style(k, bold=True, bg=WHITE)
        ws.column_dimensions["A"].width = 24
        v = ws.cell(row=row, column=2, value=value)
        style(v, bg=GREEN, num_fmt=num_fmt)
        ws.column_dimensions["B"].width = 18

    def col_hdr(cell, text):
        cell.value = text
        style(cell, bold=True, color="FFFFFF", bg=BLUE, halign="center")

    def inp(cell, value, num_fmt=None):
        cell.value = value
        style(cell, bg=GREEN, halign="center", num_fmt=num_fmt)

    def row_lbl(cell, text):
        cell.value = text
        style(cell, bold=True, bg=WHITE)

    # ── Default matrices ───────────────────────────────────────────────────
    def _default_matrix(n):
        m = np.zeros((n, n))
        if n == 3:
            m[0] = [0.80, 0.15, 0.05]
            m[1] = [0.00, 0.80, 0.20]
            m[2] = [0.00, 0.00, 1.00]
        else:
            for i in range(n - 1):
                m[i, i]   = 0.90
                m[i, i+1] = 0.10
            m[-1, -1] = 1.0
        return m

    std_mat  = _default_matrix(n_states)
    ni_mat   = std_mat.copy()
    std_util = np.linspace(1.0, 0.0, n_states).tolist()
    ni_util  = std_util.copy()

    # ─────────────────────────────────────────────────────────────────────
    # SHEET 1 · Settings
    # ─────────────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Settings"
    ws.sheet_view.showGridLines = False

    title_row(ws, 1, "CEA N-State Markov Model — Configuration Template", 2)
    inst_row(ws, 2,
             "Fill all green cells. Do NOT rename or reorder sheet tabs. "
             "Upload this file to the app to auto-load all parameters.", 2)

    section_row(ws, 3, "Global Parameters", 2)
    kv(ws, 4, "n_states",      n_states)
    kv(ws, 5, "n_cycles",      20)
    kv(ws, 6, "wtp",           50000,  num_fmt='$#,##0')
    kv(ws, 7, "discount_rate", 0.03,   num_fmt='0.00%')

    section_row(ws, 8, "State Names  (row 0 = starting / initial state)", 2)
    inst_row(ws, 9,
             "List every state in order. The first state is where all patients start.", 2)
    for i, name in enumerate(state_names, start=10):
        kv(ws, i, f"state_{i - 9}", name)

    # ─────────────────────────────────────────────────────────────────────
    # HELPER — add a Transition Matrix sheet
    # ─────────────────────────────────────────────────────────────────────
    def add_trans(sheet_name, matrix, label):
        ws2 = wb.create_sheet(sheet_name)
        ws2.sheet_view.showGridLines = False
        nc = n_states + 1

        title_row(ws2, 1, f"{label} — Transition Probability Matrix", nc)
        inst_row(ws2, 2,
                 "Each row = FROM state.  Each column = TO state.  "
                 "Every row must sum exactly to 1.0", nc)
        section_row(ws2, 3, "Rows: From state  |  Columns: To state", nc)

        # Corner
        c = ws2.cell(row=4, column=1, value="From \\ To")
        style(c, bold=True, color="FFFFFF", bg=BLUE, halign="center")
        ws2.column_dimensions["A"].width = 16

        for j, sn in enumerate(state_names, start=2):
            col_hdr(ws2.cell(row=4, column=j), sn)
            ws2.column_dimensions[get_column_letter(j)].width = 14

        for i, from_name in enumerate(state_names):
            row_lbl(ws2.cell(row=5+i, column=1), from_name)
            for j in range(n_states):
                inp(ws2.cell(row=5+i, column=j+2),
                    round(float(matrix[i, j]), 6), num_fmt="0.000000")

    # ─────────────────────────────────────────────────────────────────────
    # HELPER — add a Utilities sheet
    # ─────────────────────────────────────────────────────────────────────
    def add_util(sheet_name, utils, label):
        ws2 = wb.create_sheet(sheet_name)
        ws2.sheet_view.showGridLines = False

        title_row(ws2, 1, f"{label} — State Utilities (QALYs per cycle)", 2)
        inst_row(ws2, 2,
                 "0.0 = worst health (Dead)   1.0 = perfect health (Well)", 2)
        col_hdr(ws2.cell(row=3, column=1), "State")
        col_hdr(ws2.cell(row=3, column=2), "Utility (0 – 1)")
        ws2.column_dimensions["A"].width = 18
        ws2.column_dimensions["B"].width = 16

        for i, (sn, u) in enumerate(zip(state_names, utils), start=4):
            row_lbl(ws2.cell(row=i, column=1), sn)
            inp(ws2.cell(row=i, column=2), round(u, 4), num_fmt="0.0000")

    # ─────────────────────────────────────────────────────────────────────
    # HELPER — add a Costs sheet
    # ─────────────────────────────────────────────────────────────────────
    def add_cost(sheet_name, rows_data, label):
        ws2 = wb.create_sheet(sheet_name)
        ws2.sheet_view.showGridLines = False

        COLS   = ["State", "Subgroup", "Item", "Cost ($/year)",
                  "Distribution", "SE/SD"]
        WIDTHS = [18, 14, 22, 14, 16, 10]

        title_row(ws2, 1, f"{label} — Annual Cost Line Items", len(COLS))
        inst_row(ws2, 2,
                 "Subgroup: Medical | Non-Medical | Indirect   "
                 "Distribution: Gamma | Lognormal | Normal | Uniform | "
                 "Triangular | Beta | Fixed   "
                 "SE/SD: standard error (used in PSA; set 0 for fixed)", len(COLS))

        for j, (col, w) in enumerate(zip(COLS, WIDTHS), start=1):
            col_hdr(ws2.cell(row=3, column=j), col)
            ws2.column_dimensions[get_column_letter(j)].width = w

        s1 = state_names[0]
        s2 = state_names[min(1, n_states - 1)]

        for i, rd in enumerate(rows_data, start=4):
            for j, v in enumerate(rd, start=1):
                c = ws2.cell(row=i, column=j)
                c.value = v
                style(c, bg=GREEN,
                      num_fmt='$#,##0.00' if j == 4 else
                              '0.00'      if j == 6 else None)

        # Extra empty rows so user can add more rows
        for extra in range(4 + len(rows_data), 4 + len(rows_data) + 5):
            for j in range(1, len(COLS) + 1):
                c = ws2.cell(row=extra, column=j)
                style(c, bg=GREEN)

        return ws2

    # ─────────────────────────────────────────────────────────────────────
    # Build the 6 data sheets
    # ─────────────────────────────────────────────────────────────────────
    s1 = state_names[0]
    s2 = state_names[min(1, n_states - 1)]

    add_trans("Std_TransMatrix", std_mat,  "Standard Care")
    add_util ("Std_Utilities",   std_util, "Standard Care")
    add_cost ("Std_Costs", [
        [s1, "Medical",     "Annual Checkup",         1000.0,  "Gamma", 100.0],
        [s2, "Medical",     "Hospital Stay",           5000.0,  "Gamma", 500.0],
        [s2, "Non-Medical", "Patient Transport",        300.0,  "Gamma",  30.0],
    ], "Standard Care")

    add_trans("NI_TransMatrix", ni_mat,   "New Intervention")
    add_util ("NI_Utilities",   ni_util,  "New Intervention")
    add_cost ("NI_Costs", [
        [s1, "Medical",     "Annual Checkup",         1000.0,  "Gamma", 100.0],
        [s1, "Medical",     "New Drug (annual)",      3500.0,  "Gamma", 350.0],
        [s2, "Medical",     "Hospital Stay",           5000.0,  "Gamma", 500.0],
        [s2, "Non-Medical", "Patient Transport",        300.0,  "Gamma",  30.0],
    ], "New Intervention")

    wb.save(target)


# ─────────────────────────────────────────────────────────────────────────────
# TEMPLATE PARSING
# ─────────────────────────────────────────────────────────────────────────────

def parse_template(file) -> tuple:
    """
    Reads a filled-in CEA template (.xlsx) and returns:
        (params_dict, None)      on success
        (None, error_message)    on failure

    params_dict keys:
        n_states, state_names, n_cycles, wtp, discount_rate,
        std_p_df, std_u_df, std_cost_df,
        new_p_df, new_u_df, new_cost_df
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file, data_only=True)

        required = [
            "Settings",
            "Std_TransMatrix", "Std_Utilities", "Std_Costs",
            "NI_TransMatrix",  "NI_Utilities",  "NI_Costs",
        ]
        missing = [s for s in required if s not in wb.sheetnames]
        if missing:
            return None, f"Missing sheet(s): {', '.join(missing)}"

        # ── Settings ──────────────────────────────────────────────────────
        ws_s = wb["Settings"]
        kv = {}
        for row in ws_s.iter_rows(values_only=True):
            k, v = row[0], row[1]
            if k and v is not None and str(k).strip():
                kv[str(k).strip()] = v

        n_states = max(2, min(10, int(kv.get("n_states", 3))))
        n_cycles = max(1,  min(50, int(kv.get("n_cycles", 20))))
        wtp      = max(1000.0, float(kv.get("wtp", 50000)))
        dr_raw   = float(kv.get("discount_rate", 0.03))
        # Snap to nearest 0.01 and clamp [0, 0.10]
        discount_rate = round(max(0.0, min(0.10, round(dr_raw * 100) / 100)), 2)

        state_names = [
            str(kv.get(f"state_{i+1}", f"State {i+1}")).strip()
            for i in range(n_states)
        ]

        # ── Transition matrix ──────────────────────────────────────────────
        # Data rows start at row 5 (title=1, inst=2, section=3, header=4)
        def parse_matrix(sheet_name):
            ws2 = wb[sheet_name]
            mat = np.zeros((n_states, n_states))
            for i in range(n_states):
                for j in range(n_states):
                    v = ws2.cell(row=5+i, column=2+j).value
                    mat[i, j] = float(v) if v is not None else 0.0
            return mat

        # ── Utilities ─────────────────────────────────────────────────────
        # Data rows start at row 4 (title=1, inst=2, header=3)
        def parse_utilities(sheet_name):
            ws2 = wb[sheet_name]
            utils = []
            for i in range(n_states):
                v = ws2.cell(row=4+i, column=2).value
                utils.append(float(v) if v is not None else 0.0)
            return utils

        # ── Costs ─────────────────────────────────────────────────────────
        # Data rows start at row 4
        def parse_costs(sheet_name):
            ws2 = wb[sheet_name]
            rows = []
            for row in ws2.iter_rows(min_row=4, values_only=True):
                if row[0] is None or str(row[0]).strip() == "":
                    continue
                rows.append({
                    "State":         str(row[0]).strip(),
                    "Subgroup":      str(row[1]).strip() if row[1] else "Medical",
                    "Item":          str(row[2]).strip() if row[2] else "",
                    "Cost ($/year)": float(row[3]) if row[3] is not None else 0.0,
                    "Distribution":  str(row[4]).strip() if row[4] else "Gamma",
                    "SE/SD":         float(row[5]) if row[5] is not None else 0.0,
                })
            if rows:
                return pd.DataFrame(rows)
            return pd.DataFrame(columns=[
                "State", "Subgroup", "Item",
                "Cost ($/year)", "Distribution", "SE/SD"
            ])

        # ── Assemble ──────────────────────────────────────────────────────
        std_mat  = parse_matrix("Std_TransMatrix")
        ni_mat   = parse_matrix("NI_TransMatrix")
        std_util = parse_utilities("Std_Utilities")
        ni_util  = parse_utilities("NI_Utilities")
        std_cost = parse_costs("Std_Costs")
        ni_cost  = parse_costs("NI_Costs")

        std_p_df = pd.DataFrame(std_mat, columns=state_names, index=state_names)
        ni_p_df  = pd.DataFrame(ni_mat,  columns=state_names, index=state_names)
        std_u_df = pd.DataFrame({"State": state_names, "Utility (0-1)": std_util})
        ni_u_df  = pd.DataFrame({"State": state_names, "Utility (0-1)": ni_util})

        return {
            "n_states":      n_states,
            "state_names":   state_names,
            "n_cycles":      n_cycles,
            "wtp":           int(wtp),
            "discount_rate": discount_rate,
            "std_p_df":      std_p_df,
            "std_u_df":      std_u_df,
            "std_cost_df":   std_cost,
            "new_p_df":      ni_p_df,
            "new_u_df":      ni_u_df,
            "new_cost_df":   ni_cost,
        }, None

    except Exception as exc:
        return None, str(exc)


# ─────────────────────────────────────────────────────────────────────────────
# CLI — generate blank template
# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import sys
    out = sys.argv[1] if len(sys.argv) > 1 else "CEA_Model_Template.xlsx"
    _write_template(out)
    print(f"Template written -> {out}")
